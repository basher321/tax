"""Excel export for certificate correction (item 11).

Exports the underlying Transaction rows for a set of certificates in the
exact Depot-SCB import schema — minus Sum of VDS — so the corrected file
round-trips unchanged through import_depot_workbook (item 2). Cells touched
by an anomaly (per services.validation.check_certificate) are highlighted.
"""
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from sqlalchemy.orm import Session

from ..models.entities import Certificate, Transaction
from .certificate_generator import get_org_settings
from .validation import check_certificate

# Depot-SCB schema, minus Sum of VDS — must exactly match what
# excel_import.import_depot_workbook accepts.
EXPORT_COLUMNS = [
    ("Category", "category"),
    ("Cheque Date", "cheque_date"),
    ("Cheque Number", "cheque_number"),
    ("Supplier Name", "supplier_name"),
    ("Supplier Address", "supplier_address"),
    ("Bank Name", "bank_name"),
    ("WhatsApp No.", "whatsapp_no"),
    ("Email", "email"),
    ("Depot Code", "depot_code"),
    ("Description of Payment", "description_of_payment"),
    ("Sum of Bill Amount", "sum_of_bill_amount"),
    ("Sum of TDS", "sum_of_tds"),
    ("Match", "match"),
    ("Section", "section"),
    ("TIN", "tin"),
    ("Challan No", "challan_no"),
    ("Challan Date", "challan_date"),
    ("Cheque/Challan SL", "cheque_challan_sl"),
    ("Month", "month"),
    ("Total Challan Amount", "total_challan_amount"),
    ("Remarks", "remarks"),
]

# Anomaly code -> export column header(s) to highlight for that certificate's rows.
ANOMALY_COLUMN_MAP = {
    "MISSING_EMAIL": ["Email"],
    "INVALID_EMAIL": ["Email"],
    "MISSING_WHATSAPP": ["WhatsApp No."],
    "INVALID_WHATSAPP": ["WhatsApp No."],
    "MISSING_TIN": ["TIN"],
    "MISSING_CHALLAN": ["Challan No", "Challan Date"],
    "TDS_MISMATCH": ["Sum of TDS"],
}

_HIGHLIGHT = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


def export_certificates_to_excel(db: Session, certificates: list[Certificate]) -> Workbook:
    """One row per underlying Transaction for the given certificates."""
    org = get_org_settings(db)
    wb = Workbook()
    ws = wb.active
    ws.title = "Depot-SCB"
    ws.append([header for header, _ in EXPORT_COLUMNS])

    for cert in certificates:
        anomalies = check_certificate(db, cert, org)
        highlight_cols = {
            col for a in anomalies for col in ANOMALY_COLUMN_MAP.get(a.code, [])
        }

        txns = (
            db.query(Transaction)
            .filter(Transaction.company_id == cert.company_id, Transaction.tin == cert.tin,
                    Transaction.fiscal_year == cert.period)
            .order_by(Transaction.cheque_date, Transaction.id)
            .all()
        )
        for t in txns:
            ws.append([getattr(t, attr) for _, attr in EXPORT_COLUMNS])
            if highlight_cols:
                row_idx = ws.max_row
                for col_idx, (header, _) in enumerate(EXPORT_COLUMNS, start=1):
                    if header in highlight_cols:
                        ws.cell(row=row_idx, column=col_idx).fill = _HIGHLIGHT

    return wb
