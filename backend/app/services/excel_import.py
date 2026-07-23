"""Excel import service for the Depot-SCB sheet.

Parses the workbook directly with openpyxl — no Excel formula evaluation
happens at runtime. Every row is accepted and stored as-is: no required-field,
TIN-format, or number/date-parse checks reject or skip a row. Values that
don't parse (e.g. an unparseable date) are simply stored as null.
"""
import math
import re
from datetime import date, datetime

import openpyxl
from sqlalchemy.orm import Session, joinedload

from ..models.entities import (
    ContactKind,
    ImportBatch,
    Supplier,
    SupplierContact,
    Transaction,
)

# The source columns, exactly as named in the sheet, mapped to DB fields.
# ("Challan No" appears without a trailing dot in the file; we accept both.
# "Payment Date" is the current template header; "Cheque Date" is accepted
# too for older workbooks — both map to the same DB column.)
SOURCE_COLUMNS = {
    "Category": "category",
    "Payment Date": "cheque_date",
    "Cheque Date": "cheque_date",
    "Cheque Number": "cheque_number",
    "Supplier Name": "supplier_name",
    "Supplier Address": "supplier_address",
    "Bank Name": "bank_name",
    "WhatsApp No.": "whatsapp_no",
    "Email": "email",
    "Depot Code": "depot_code",
    "Description of Payment": "description_of_payment",
    "Sum of Bill Amount": "sum_of_bill_amount",
    "Sum of TDS": "sum_of_tds",
    # Sum of VDS is intentionally excluded from import/field mapping/storage.
    # Base Amount / TDS Rate are optional: when both are present, TDS is
    # computed as Base Amount x TDS Rate instead of taken literally.
    "Base Amount": "base_amount",
    "TDS Rate": "tds_rate",
    "Match": "match",
    "Section": "section",
    "TIN": "tin",
    "Challan No": "challan_no",
    "Challan No.": "challan_no",
    "Challan Date": "challan_date",
    "Cheque/Challan SL": "cheque_challan_sl",
    "Month": "month",
    "Total Challan Amount": "total_challan_amount",
    "Remarks": "remarks",
}

# Used only to locate which row is the header row in the sheet — not a
# validation of the data itself. A sheet missing either of these has no
# usable header to anchor on at all.
HEADER_ANCHOR_FIELDS = ["Supplier Name", "TIN"]

MONTH_RE = re.compile(r"^([A-Za-z]+)'(\d{2})$")
_MONTHS = {m.lower(): i for i, m in enumerate(
    ["January", "February", "March", "April", "May", "June", "July",
     "August", "September", "October", "November", "December"], start=1)}


def parse_month_label(label: str) -> date | None:
    """'December'25' -> date(2025, 12, 1)."""
    if not label:
        return None
    m = MONTH_RE.match(str(label).strip())
    if not m:
        return None
    name, yy = m.groups()
    mo = _MONTHS.get(name.lower())
    if not mo:
        return None
    return date(2000 + int(yy), mo, 1)


def fiscal_year_for(d: date) -> str:
    """Bangladeshi fiscal year (Jul 1 – Jun 30), e.g. 2025-12 -> '2025-26'."""
    start = d.year if d.month >= 7 else d.year - 1
    return f"{start}-{str(start + 1)[-2:]}"


def _parse_date(value) -> date | None:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_number(value) -> float | None:
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    try:
        f = float(value)
        return None if math.isnan(f) else f
    except (TypeError, ValueError):
        return None


def _clean(value):
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    s = str(value).strip()
    return s or None


def _find_header_row(ws) -> int | None:
    """Locate the 0-based row index (within the first 10 rows) that looks like
    the real header — i.e. contains every HEADER_ANCHOR_FIELDS column name.
    Sheets aren't guaranteed to include every SOURCE_COLUMNS field, so
    anchoring on a couple of fields every workbook necessarily has is what
    actually generalizes across differently-trimmed templates."""
    required = set(HEADER_ANCHOR_FIELDS)
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True)):
        vals = {str(v).strip() if v is not None else "" for v in row}
        if required <= vals:
            return i
    return None


def load_depot_sheet(path: str, sheet_name: str = "Depot-SCB") -> tuple[list[str], list[dict]]:
    """Read the Depot sheet, locating the header row that contains the
    HEADER_ANCHOR_FIELDS column names.
    Returns (columns, rows); each row is a dict of column name -> cell value,
    plus a 1-based '__excel_row' Excel row number."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        sheet_names = wb.sheetnames
        candidates = [sheet_name] if sheet_name in sheet_names else []
        candidates.extend([name for name in sheet_names if name not in candidates])

        selected_sheet = None
        header_idx = None
        for candidate in candidates:
            header_idx = _find_header_row(wb[candidate])
            if header_idx is not None:
                selected_sheet = candidate
                break

        if header_idx is None:
            raise ValueError(
                "Could not find a header row containing: "
                + ", ".join(HEADER_ANCHOR_FIELDS)
            )

        ws = wb[selected_sheet]
        header_row_num = header_idx + 1  # 1-based
        header_values = next(
            ws.iter_rows(min_row=header_row_num, max_row=header_row_num, values_only=True)
        )
        columns = [str(c).strip() if c is not None else "" for c in header_values]

        rows = []
        for excel_row_num, values in enumerate(
            ws.iter_rows(min_row=header_row_num + 1, values_only=True), start=header_row_num + 1
        ):
            # Skip fully blank rows — common trailing padding from a
            # formatted-but-empty range (e.g. borders/column formatting
            # applied far past the real data), not a data row.
            if all(v is None for v in values):
                continue
            row = {columns[i]: (values[i] if i < len(values) else None) for i in range(len(columns))}
            row["__excel_row"] = excel_row_num
            rows.append(row)
        return columns, rows
    finally:
        wb.close()


def _get_or_create_supplier(db: Session, cache: dict[str, Supplier], company_id: int,
                             tin: str, name: str, address, email, wa) -> Supplier:
    """`cache` holds every supplier already seen for this company — prefetched
    once by the caller, plus every supplier created earlier in this same
    import — so a multi-thousand-row workbook doesn't cost a supplier lookup
    query per row (most rows share a relative handful of distinct suppliers)."""
    sup = cache.get(tin)
    if not sup:
        sup = Supplier(company_id=company_id, tin=tin, name=name, address=address)
        db.add(sup)
        db.flush()  # need the new supplier's id before Transaction.supplier_id can reference it
        cache[tin] = sup
    else:
        # Keep the most recent non-empty name/address.
        if name:
            sup.name = name
        if address:
            sup.address = address
    for kind, value in ((ContactKind.EMAIL, email), (ContactKind.WHATSAPP, wa)):
        if value:
            exists = any(c.kind == kind and c.value == value for c in sup.contacts)
            if not exists:
                sup.contacts.append(SupplierContact(kind=kind, value=value))
    return sup


def import_depot_workbook(db: Session, path: str, filename: str, company_id: int) -> ImportBatch:
    """Import every row of the Depot-SCB sheet as-is — no validation, no
    skipped/error rows. Missing or unparseable values are stored as null."""
    columns, rows = load_depot_sheet(path)
    batch = ImportBatch(company_id=company_id, filename=filename, kind="depot", total_rows=len(rows))
    db.add(batch)
    db.flush()

    # One query for every supplier this company already has, instead of one
    # per row — the dominant cost on large imports otherwise.
    supplier_cache = {
        s.tin: s for s in
        db.query(Supplier).options(joinedload(Supplier.contacts))
          .filter(Supplier.company_id == company_id)
    }

    for row in rows:
        def val(col):
            return row.get(col)

        tin = _clean(val("TIN"))
        if tin:
            tin = re.sub(r"\D", "", tin) or tin

        base_amount = _parse_number(val("Base Amount"))
        tds_rate = _parse_number(val("TDS Rate"))
        sum_of_tds = _parse_number(val("Sum of TDS"))
        # Base Amount x TDS Rate overrides the literal Sum of TDS column when
        # both are present and parse as numbers; otherwise Sum of TDS (already
        # based on Sum of Bill Amount upstream in the sheet) is used as-is.
        computed_tds = round(base_amount * tds_rate, 2) if (base_amount is not None and tds_rate is not None) else sum_of_tds

        cheque_date = _parse_date(val("Payment Date") if "Payment Date" in columns else val("Cheque Date"))
        challan_date = _parse_date(val("Challan Date"))
        month_label = _clean(val("Month"))
        month_date = parse_month_label(month_label) if month_label else None

        supplier = None
        if tin:
            supplier = _get_or_create_supplier(
                db, supplier_cache, company_id, tin, _clean(val("Supplier Name")),
                _clean(val("Supplier Address")), _clean(val("Email")),
                _clean(val("WhatsApp No.")),
            )

        basis = month_date or cheque_date
        txn = Transaction(
            company_id=company_id,
            batch_id=batch.id,
            supplier_id=supplier.id if supplier else None,
            category=_clean(val("Category")),
            cheque_date=cheque_date,
            cheque_number=_clean(val("Cheque Number")),
            supplier_name=_clean(val("Supplier Name")),
            supplier_address=_clean(val("Supplier Address")),
            bank_name=_clean(val("Bank Name")),
            whatsapp_no=_clean(val("WhatsApp No.")),
            email=_clean(val("Email")),
            depot_code=_clean(val("Depot Code")),
            description_of_payment=_clean(val("Description of Payment")),
            sum_of_bill_amount=_parse_number(val("Sum of Bill Amount")),
            sum_of_tds=computed_tds,
            base_amount=base_amount,
            tds_rate=tds_rate,
            match=_clean(val("Match")),
            section=_clean(val("Section")),
            tin=tin,
            challan_no=_clean(val("Challan No") if "Challan No" in columns
                              else val("Challan No.")),
            challan_date=challan_date,
            cheque_challan_sl=_clean(val("Cheque/Challan SL")),
            month=month_label,
            total_challan_amount=_parse_number(val("Total Challan Amount")),
            remarks=_clean(val("Remarks")),
            fiscal_year=fiscal_year_for(basis) if basis else None,
        )
        db.add(txn)

    batch.ok_rows = len(rows)
    batch.error_rows = 0
    db.commit()
    return batch
