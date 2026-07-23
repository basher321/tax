"""Tests: Excel import parsing (incl. Base Amount x TDS Rate + VDS exclusion),
aggregation math, concurrent number allocation, anomaly detection, bulk
anomaly check / bulk send, per-company letterhead resolution, multi-company
isolation, and offline dispatch queue."""
import io
import os
import smtplib
import threading
from datetime import date
from types import SimpleNamespace

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook
from pydantic import ValidationError
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

import app.models  # noqa: F401 — register tables
from app.api.routes import (
    bulk_anomaly_check,
    bulk_dispatch,
    onboard_supplier,
)
from app.database import Base
from app.models.entities import (
    Certificate,
    Company,
    ContactKind,
    DispatchChannel,
    DispatchStatus,
    DispatchJob,
    PartyType,
    Signature,
    Supplier,
    SupplierContact,
    TaxRate,
    Transaction,
)
from app.schemas import BulkDispatchRequest, BulkFilterRequest, SupplierCreate
from app.services import rate_hook
from app.services.aggregation import build_certificate_data, list_groupings
from app.services.amount_in_words import amount_in_words
from app.services.certificate_generator import (
    GenerationError,
    generate_certificate,
    get_org_settings,
)
from app.services.dispatch.queue import DispatchBlocked, enqueue_dispatch, process_queue
from app.services.dispatch.email_sender import send_test_email
from app.services.dispatch.whatsapp_sender import send_certificate_whatsapp
from app.services.excel_import import (
    fiscal_year_for,
    import_depot_workbook,
    parse_month_label,
)
from app.services.numbering import allocate_certificate_number, get_numbering_config
from app.services.pdf_renderer import render_certificate_pdf
from app.services.validation import check_certificate

_TINY_PNG = (
    b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01"
    b"\x08\x06\x00\x00\x00\x1f\x15\xc4\x89\x00\x00\x00\nIDATx\x9cc\x00\x01"
    b"\x00\x00\x05\x00\x01\r\n-\xb4\x00\x00\x00\x00IEND\xaeB`\x82"
)

HEADERS = [
    "Category", "Cheque Date", "Cheque Number", "Supplier Name",
    "Supplier Address", "Bank Name", "WhatsApp No.", "Email", "Depot Code",
    "Description of Payment", "Sum of Bill Amount", "Sum of TDS", "Sum of VDS",
    "Base Amount", "TDS Rate",
    "Match", "Section", "TIN", "Challan No", "Challan Date",
    "Cheque/Challan SL", "Month", "Total Challan Amount", "Remarks",
]


@pytest.fixture
def db(tmp_path):
    # File-backed SQLite so the concurrency test can open multiple sessions.
    url = f"sqlite:///{tmp_path}/test.db"
    engine = create_engine(url, connect_args={"check_same_thread": False})
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine, expire_on_commit=False)
    s = Session()
    yield s
    s.close()


@pytest.fixture
def company(db):
    c = Company(name="Renata PLC", is_default=True)
    db.add(c)
    db.commit()
    return c


@pytest.fixture
def company2(db):
    c = Company(name="Second Co")
    db.add(c)
    db.commit()
    return c


@pytest.fixture
def sample_xlsx(tmp_path):
    """Known sample: 3 rows for TIN A (Dec'25), 1 bad row, 1 row for TIN B."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Depot-SCB"
    ws.append([None] * len(HEADERS))
    ws.append([None] * len(HEADERS))
    ws.append(HEADERS)

    def row(name, tin, bill, tds, month, cheque=date(2025, 12, 1),
            section="89", challan="2526-001", chdate="15/01/2026",
            total_ch=100000, email=None, wa=None, base_amount=None, tds_rate=None):
        ws.append(["Depot", cheque, "CHQ1", name, None, None, wa, email, None,
                   None, bill, tds, -bill * 0.05 if bill else None,
                   base_amount, tds_rate, None,
                   section, tin, challan, chdate, 1, month, total_ch, None])

    row("Alpha Traders", "111122223333", 100000, 5000, "December'25")
    row("Alpha Traders", "111122223333", 50000, 2500, "January'26",
        cheque=date(2026, 1, 10))
    row("Alpha Traders", "111122223333", 20000, 1000, "January'26",
        cheque=date(2026, 1, 20))
    row("Beta Supplies", "444455556666", 80000, 4000, "December'25",
        email="beta@example.com", wa="+8801711111111")
    row("Bad Row Ltd", "12AB", 10000, 500, "December'25")  # invalid TIN

    path = tmp_path / "sample.xlsx"
    wb.save(path)
    return str(path)


# ------------------------------------------------------------- Import -------
def test_import_parses_and_validates(db, company, sample_xlsx):
    batch = import_depot_workbook(db, sample_xlsx, "sample.xlsx", company.id)
    assert batch.company_id == company.id
    assert batch.total_rows == 5
    assert batch.ok_rows == 4
    assert batch.error_rows == 1
    msgs = [e.message for e in batch.errors]
    assert any("not a valid 12-digit TIN" in m for m in msgs)
    # All fields land on the transaction row.
    t = db.query(Transaction).filter_by(tin="111122223333").first()
    assert t.company_id == company.id
    assert t.category == "Depot" and t.section == "89"
    assert t.challan_no == "2526-001" and t.challan_date == date(2026, 1, 15)
    assert t.month == "December'25" and t.fiscal_year == "2025-26"


def test_vds_never_populated_by_import(db, company, sample_xlsx):
    """Sum of VDS is present in the sheet but must never be written (item 2)."""
    import_depot_workbook(db, sample_xlsx, "sample.xlsx", company.id)
    txns = db.query(Transaction).filter_by(tin="111122223333").all()
    assert txns and all(t.sum_of_vds is None for t in txns)


def test_base_amount_times_tds_rate_overrides_sum_of_tds(db, company, tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Depot-SCB"
    ws.append(HEADERS)
    # Sum of TDS says 999 (should be ignored) because Base Amount x TDS Rate is present.
    ws.append(["Depot", date(2025, 12, 1), "CHQ9", "Gamma Traders", None, None,
               None, None, None, None, 100000, 999, None,
               50000, 0.1,  # Base Amount x TDS Rate = 5000
               None, "89", "999900001111", "2526-777", "15/01/2026",
               1, "December'25", 100000, None])
    path = tmp_path / "base_rate.xlsx"
    wb.save(path)

    import_depot_workbook(db, str(path), "base_rate.xlsx", company.id)
    t = db.query(Transaction).filter_by(tin="999900001111").first()
    assert t.sum_of_tds == 5000  # 50000 * 0.1, not the literal 999


def test_month_and_fiscal_year_parsing():
    assert parse_month_label("December'25") == date(2025, 12, 1)
    assert fiscal_year_for(date(2025, 12, 5)) == "2025-26"
    assert fiscal_year_for(date(2026, 3, 1)) == "2025-26"
    assert fiscal_year_for(date(2026, 7, 1)) == "2026-27"


# -------------------------------------------------------- Aggregation -------
def test_aggregation_math_known_sample(db, company, sample_xlsx):
    import_depot_workbook(db, sample_xlsx, "sample.xlsx", company.id)
    data = build_certificate_data(db, company.id, "111122223333", "2025-26")
    # SUM logic: 100000+50000+20000 payment; 5000+2500+1000 TDS.
    assert data.total_payment == 170000
    assert data.total_tax_deducted == 8500
    assert data.total_challan_related == 8500  # Section 07 total == TDS total
    assert [ln.sl for ln in data.lines] == [1, 2, 3]  # ordered by date
    assert data.period_from == date(2025, 7, 1)
    assert data.period_to == date(2026, 1, 31)  # end of latest covered month


def test_amount_in_words_matches_template_style():
    assert amount_in_words(41382) == \
        "Forty One Thousand Three Hundred Eighty Two Only."


# ------------------------------------------------- Generation rules ---------
def test_supplier_only_enforced_at_service(db, company, sample_xlsx):
    import_depot_workbook(db, sample_xlsx, "sample.xlsx", company.id)
    sup = db.query(Supplier).filter_by(tin="111122223333").first()
    sup.party_type = PartyType.EMPLOYEE
    db.commit()
    with pytest.raises(GenerationError, match="supplier-only"):
        generate_certificate(db, company.id, "111122223333", "2025-26")


def test_duplicate_certificate_blocked(db, company, sample_xlsx):
    import_depot_workbook(db, sample_xlsx, "sample.xlsx", company.id)
    generate_certificate(db, company.id, "444455556666", "2025-26")
    with pytest.raises(GenerationError, match="already exists"):
        generate_certificate(db, company.id, "444455556666", "2025-26")


# ------------------------------------------- Multi-company isolation --------
def test_two_companies_same_tin_isolated(db, company, company2, sample_xlsx):
    """The same TIN can independently onboard/transact/generate under two
    different companies without colliding (full multi-tenant scope)."""
    import_depot_workbook(db, sample_xlsx, "sample.xlsx", company.id)
    import_depot_workbook(db, sample_xlsx, "sample.xlsx", company2.id)

    suppliers = db.query(Supplier).filter_by(tin="111122223333").all()
    assert len(suppliers) == 2
    assert {s.company_id for s in suppliers} == {company.id, company2.id}

    cert1 = generate_certificate(db, company.id, "444455556666", "2025-26")
    cert2 = generate_certificate(db, company2.id, "444455556666", "2025-26")
    assert cert1.id != cert2.id
    assert cert1.certificate_no != cert2.certificate_no  # independent sequences


def test_supplier_onboarding_upserts_per_company_not_globally(db, company, company2):
    body1 = SupplierCreate(company_id=company.id, name="Gamma", address="Dhaka",
                           tin="555566667777", bin="B1", email="a@b.com",
                           whatsapp="+8801711111111")
    body2 = SupplierCreate(company_id=company2.id, name="Gamma Co2", address="Ctg",
                           tin="555566667777", bin="B2", email="c@d.com",
                           whatsapp="+8801711112222")
    sup1 = onboard_supplier(body1, db)
    sup2 = onboard_supplier(body2, db)
    assert sup1.id != sup2.id
    assert db.query(Supplier).filter_by(tin="555566667777").count() == 2


# --------------------------------------------- Numbering concurrency --------
def test_concurrent_number_allocation(tmp_path):
    url = f"sqlite:///{tmp_path}/conc.db"
    engine = create_engine(url, connect_args={"check_same_thread": False,
                                              "timeout": 30})
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine, expire_on_commit=False)

    s0 = Session()
    company = Company(name="ACME", is_default=True)
    s0.add(company)
    s0.commit()
    cfg = get_numbering_config(s0, company.id)
    cfg.company_token = "ACME"
    cfg.pad_width = 5
    s0.commit()
    company_id = company.id
    s0.close()

    results, errors = [], []

    def worker():
        s = Session()
        try:
            n = allocate_certificate_number(s, company_id, "2025-26")
            s.commit()
            results.append(n)
        except Exception as e:  # noqa: BLE001
            s.rollback()
            errors.append(e)
        finally:
            s.close()

    threads = [threading.Thread(target=worker) for _ in range(20)]
    for t in threads:
        t.start()
    for t in threads:
        t.join()

    assert not errors, errors
    assert len(results) == 20
    assert len(set(results)) == 20, "numbers must be unique under concurrency"
    suffixes = sorted(int(r.split("/")[-1]) for r in results)
    assert suffixes == list(range(1, 21)), "numbers must be sequential"
    assert results[0].startswith("ACME/2025-26/")


# ---------------------------------------------------- Anomaly checks --------
def test_anomaly_detection_rules(db, company, sample_xlsx):
    import_depot_workbook(db, sample_xlsx, "sample.xlsx", company.id)
    cert = generate_certificate(db, company.id, "111122223333", "2025-26")  # no contacts
    org = get_org_settings(db)
    codes = {a.code for a in check_certificate(db, cert, org)}
    assert "MISSING_EMAIL" in codes
    assert "MISSING_WHATSAPP" in codes
    assert "MISSING_SEAL" in codes  # no seal uploaded in this test env

    # TDS reconciliation: expected rate 5%, actual is 5% -> no mismatch...
    db.add(TaxRate(section="89", kind="tds", rate=0.05,
                   effective_from=date(2025, 7, 1)))
    db.commit()
    codes = {a.code for a in check_certificate(db, cert, org)}
    assert "TDS_MISMATCH" not in codes
    # ...but a newer expected rate (10%) makes the 5% deduction anomalous.
    db.add(TaxRate(section="89", kind="tds", rate=0.10,
                   effective_from=date(2026, 6, 1)))
    db.commit()
    codes = {a.code for a in check_certificate(db, cert, org)}
    assert "TDS_MISMATCH" in codes


def test_missing_challan_anomaly(db, company, sample_xlsx):
    import_depot_workbook(db, sample_xlsx, "sample.xlsx", company.id)
    for t in db.query(Transaction).filter_by(tin="444455556666"):
        t.challan_no = None
    db.commit()
    cert = generate_certificate(db, company.id, "444455556666", "2025-26")
    org = get_org_settings(db)
    codes = {a.code for a in check_certificate(db, cert, org)}
    assert "MISSING_CHALLAN" in codes


def _make_clean_certificate(db, company, sample_xlsx):
    """Beta Supplies has email+WhatsApp+challan mapped; with company seal
    filled in, its certificate should have zero anomalies."""
    import_depot_workbook(db, sample_xlsx, "sample.xlsx", company.id)
    company.seal_data = _TINY_PNG
    db.commit()
    return generate_certificate(db, company.id, "444455556666", "2025-26")


# ---------------------------------------- Bulk anomaly check (item 5) -------
def test_bulk_anomaly_check_reports_only_anomalous(db, company, sample_xlsx):
    clean = _make_clean_certificate(db, company, sample_xlsx)
    anomalous = generate_certificate(db, company.id, "111122223333", "2025-26")  # no contacts

    results = bulk_anomaly_check(BulkFilterRequest(company_id=company.id), db)
    ids = {r.certificate_id for r in results}
    assert clean.id not in ids
    assert anomalous.id in ids


# --------------------------------------------- Bulk send (item 10) ----------
def test_bulk_send_dispatches_clean_skips_anomalous(db, company, sample_xlsx, monkeypatch):
    clean = _make_clean_certificate(db, company, sample_xlsx)
    generate_certificate(db, company.id, "111122223333", "2025-26")  # stays anomalous
    get_org_settings(db).dispatch_mode = "online"
    db.commit()

    sent = []
    monkeypatch.setattr(
        "app.services.dispatch.queue.send_certificate_email",
        lambda org, cert, r, job_id=None: sent.append(r))

    results = bulk_dispatch(BulkDispatchRequest(company_id=company.id, channel="email"), db)
    by_id = {r.certificate_id: r for r in results}
    assert by_id[clean.id].ok is True
    assert sent == ["beta@example.com"]
    anomalous_result = next(r for r in results if r.certificate_id != clean.id)
    assert anomalous_result.ok is False
    assert "Skipped" in anomalous_result.error


# --------------------------------------- Per-company letterhead (item 12) ---
def test_letterhead_resolves_per_certificates_own_company(db, company, company2, sample_xlsx, tmp_path):
    header1 = _TINY_PNG
    header2 = _TINY_PNG + b"\x00"  # distinguishable bytes, still a valid PNG read path
    company.letterhead_header_data = header1
    company2.letterhead_header_data = header2
    db.commit()

    import_depot_workbook(db, sample_xlsx, "sample.xlsx", company.id)
    import_depot_workbook(db, sample_xlsx, "sample.xlsx", company2.id)
    cert1 = generate_certificate(db, company.id, "444455556666", "2025-26")
    cert2 = generate_certificate(db, company2.id, "444455556666", "2025-26")

    # Each certificate's own company_id resolves to its own letterhead —
    # never cross-contaminated, and render succeeds for both.
    assert cert1.company.letterhead_header_data == header1
    assert cert2.company.letterhead_header_data == header2
    render_certificate_pdf(db, cert1)
    render_certificate_pdf(db, cert2)
    assert cert1.pdf_data and cert2.pdf_data


# --------------------------------- Share-ready certificate image export -----
def test_certificate_image_generated_alongside_pdf(db, company, sample_xlsx):
    """Every (re)generation also rasterizes the PDF into a JPEG for
    WhatsApp/email sharing, sized to avoid WhatsApp's own re-compression
    and to stay well under typical attachment limits."""
    from PIL import Image

    import_depot_workbook(db, sample_xlsx, "sample.xlsx", company.id)
    cert = generate_certificate(db, company.id, "444455556666", "2025-26")

    assert cert.image_data
    assert 0 < len(cert.image_data) < 2 * 1024 * 1024  # under the 2 MB target

    with Image.open(io.BytesIO(cert.image_data)) as img:
        assert img.format == "JPEG"
        assert max(img.size) in range(1700, 1901)  # ~1800px long edge


def test_certificate_image_includes_overflow_pages(db, company, tmp_path):
    """A certificate with enough line items to overflow onto a second PDF
    page still carries its Seal and Signature block on whichever page the
    content ends on (normal pagination) — the exported image must include
    that page too, not silently drop it by only rasterizing page one."""
    from PIL import Image
    import pypdfium2 as pdfium

    wb = Workbook()
    ws = wb.active
    ws.title = "Depot-SCB"
    ws.append([None] * len(HEADERS))
    ws.append([None] * len(HEADERS))
    ws.append(HEADERS)
    for i in range(40):
        ws.append(["Depot", date(2025, 12, 1), f"CHQ{i}", "Bulk Traders", None, None,
                   None, None, None, None, 10000, 500, None, None, None, None,
                   "89", "999988887777", f"2526-{i:03d}", "15/01/2026",
                   1, "December'25", 100000, None])
    path = tmp_path / "bulk.xlsx"
    wb.save(path)

    import_depot_workbook(db, str(path), "bulk.xlsx", company.id)
    cert = generate_certificate(db, company.id, "999988887777", "2025-26")

    with pdfium.PdfDocument(cert.pdf_data) as pdf:
        page_count = len(pdf)
    assert page_count > 1, "test setup should force real overflow to prove the fix"

    with Image.open(io.BytesIO(cert.image_data)) as img:
        assert img.format == "JPEG"
        # Stacked multi-page image is taller than a single A4 page's
        # rendered aspect ratio (297/210 ≈ 1.41) would allow.
        assert img.height / img.width > (297 / 210) * (page_count - 0.5)


def test_certificate_image_route_self_heals_missing_image(db, company, sample_xlsx):
    """Certificates generated before the image-export feature existed have
    a PDF but no image yet — the download route rasterizes one on first
    request instead of 404ing, so the on-screen preview (which now shows
    this same image, not a separate HTML re-implementation) never breaks
    for older certificates."""
    from app.main import app
    from app.database import get_db

    import_depot_workbook(db, sample_xlsx, "sample.xlsx", company.id)
    cert = generate_certificate(db, company.id, "444455556666", "2025-26")
    assert cert.image_data  # generated normally the first time
    cert.image_data = None  # simulate a pre-existing certificate
    db.commit()

    app.dependency_overrides[get_db] = lambda: db
    try:
        client = TestClient(app)
        resp = client.get(f"/api/certificates/{cert.id}/image")
        assert resp.status_code == 200
        assert resp.headers["content-type"] == "image/jpeg"
        db.refresh(cert)
        assert cert.image_data
    finally:
        app.dependency_overrides.clear()


# ------------------------------------------------ Offline dispatch ----------
def test_offline_queue_blocks_then_overrides_then_drains(db, company, sample_xlsx, monkeypatch):
    import_depot_workbook(db, sample_xlsx, "sample.xlsx", company.id)
    cert = generate_certificate(db, company.id, "444455556666", "2025-26")
    org = get_org_settings(db)
    org.dispatch_mode = "offline"
    db.commit()

    # Anomalies (missing seal) block dispatch without an override.
    with pytest.raises(DispatchBlocked) as exc:
        enqueue_dispatch(db, cert, DispatchChannel.EMAIL, ["beta@example.com"])
    assert any(a.code == "MISSING_SEAL" for a in exc.value.anomalies)

    # Override with a logged reason -> job queued but NOT sent (offline mode).
    jobs = enqueue_dispatch(db, cert, DispatchChannel.EMAIL,
                            ["beta@example.com"],
                            override_reason="Seal pending, board approved",
                            user="admin")
    assert jobs[0].status == DispatchStatus.QUEUED

    # Simulate connectivity returning: worker drains the queue.
    sent = []
    monkeypatch.setattr(
        "app.services.dispatch.queue.send_certificate_email",
        lambda org, cert, r: sent.append(r))
    processed = process_queue(db)
    assert processed == 1
    job = db.query(DispatchJob).first()
    assert job.status == DispatchStatus.SENT
    assert sent == ["beta@example.com"]


def test_queue_retries_and_fails_after_max_attempts(db, company, sample_xlsx, monkeypatch):
    import_depot_workbook(db, sample_xlsx, "sample.xlsx", company.id)
    cert = generate_certificate(db, company.id, "444455556666", "2025-26")
    get_org_settings(db).dispatch_mode = "offline"
    db.commit()
    jobs = enqueue_dispatch(db, cert, DispatchChannel.WHATSAPP,
                            ["+8801711111111"], override_reason="ok", user="t")

    def boom(org, cert, r):
        raise RuntimeError("network down")
    monkeypatch.setattr(
        "app.services.dispatch.queue.send_certificate_whatsapp", boom)

    for _ in range(6):
        process_queue(db)
    job = db.get(DispatchJob, jobs[0].id)
    assert job.status == DispatchStatus.FAILED
    assert job.attempts == 5
    assert "network down" in job.last_error


def test_send_test_email_uses_settings_and_recipient(db, monkeypatch):
    org = get_org_settings(db)
    org.company_name = "Acme Ltd"
    org.smtp_host = "smtp.example.com"
    org.smtp_port = 587
    org.smtp_from = "tax@example.com"
    org.smtp_user = "tax@example.com"
    org.smtp_password = "app-password"
    org.smtp_use_tls = True
    db.commit()

    sent = {}

    class FakeSMTP:
        def __init__(self, host, port, timeout):
            sent["connect"] = (host, port, timeout)

        def starttls(self):
            sent["tls"] = True

        def login(self, user, password):
            sent["login"] = (user, password)

        def send_message(self, msg):
            sent["to"] = msg["To"]
            sent["from"] = msg["From"]
            sent["subject"] = msg["Subject"]

        def quit(self):
            sent["quit"] = True

    monkeypatch.setattr("smtplib.SMTP", FakeSMTP)

    # No explicit recipient -> falls back to the configured sender address.
    assert send_test_email(org) == "tax@example.com"
    assert sent["connect"] == ("smtp.example.com", 587, 30)
    assert sent["tls"] is True
    assert sent["login"] == ("tax@example.com", "app-password")
    assert sent["to"] == "tax@example.com"
    assert "tax@example.com" in sent["from"]
    assert sent["subject"] == "Tax Certificate SMTP test"


def test_send_test_email_reports_auth_failure(db, monkeypatch):
    org = get_org_settings(db)
    org.smtp_host = "smtp.example.com"
    org.smtp_from = "tax@example.com"
    org.smtp_user = "tax@example.com"
    org.smtp_password = "bad"
    db.commit()

    class FakeSMTP:
        def __init__(self, host, port, timeout):
            pass

        def starttls(self):
            pass

        def login(self, user, password):
            raise smtplib.SMTPAuthenticationError(535, b"Invalid credentials")

        def quit(self):
            pass

    monkeypatch.setattr("smtplib.SMTP", FakeSMTP)

    with pytest.raises(RuntimeError, match="SMTP authentication failed"):
        send_test_email(org, "officer@example.com")


def test_whatsapp_cloud_sends_document_without_caption_link(db, monkeypatch):
    org = get_org_settings(db)
    org.wa_provider = "cloud"
    org.wa_token = "token"
    org.wa_phone_number_id = "phone-id"
    db.commit()
    cert = Certificate(id=123, certificate_no="ACME/2025-26/1", period="2025-26")
    sent = {}

    class FakeResp:
        def raise_for_status(self):
            sent["raised"] = False

    def fake_post(url, headers=None, json=None, timeout=None):
        sent["url"] = url
        sent["headers"] = headers
        sent["json"] = json
        sent["timeout"] = timeout
        return FakeResp()

    monkeypatch.setattr("httpx.post", fake_post)
    monkeypatch.setattr(
        "app.services.dispatch.whatsapp_sender.get_settings",
        lambda: SimpleNamespace(
            public_base_url="https://tax.example.com",
            link_signing_secret="test-secret",
        ),
    )

    send_certificate_whatsapp(org, cert, "+8801711111111")

    doc = sent["json"]["document"]
    assert sent["json"]["type"] == "document"
    assert "/public/certificates/123" in doc["link"]
    assert "Download:" not in doc["caption"]
    assert "http" not in doc["caption"]
    assert doc["filename"] == "ACME_2025-26_1.pdf"


def test_whatsapp_twilio_sends_media_without_body_link(db, monkeypatch):
    org = get_org_settings(db)
    org.wa_provider = "twilio"
    org.wa_twilio_sid = "sid"
    org.wa_twilio_auth = "auth"
    org.wa_twilio_from = "+15550000000"
    db.commit()
    cert = Certificate(id=456, certificate_no="ACME/2025-26/2", period="2025-26")
    sent = {}

    class FakeResp:
        def raise_for_status(self):
            sent["raised"] = False

    def fake_post(url, auth=None, data=None, timeout=None):
        sent["url"] = url
        sent["auth"] = auth
        sent["data"] = data
        sent["timeout"] = timeout
        return FakeResp()

    monkeypatch.setattr("httpx.post", fake_post)
    monkeypatch.setattr(
        "app.services.dispatch.whatsapp_sender.get_settings",
        lambda: SimpleNamespace(
            public_base_url="https://tax.example.com",
            link_signing_secret="test-secret",
        ),
    )

    send_certificate_whatsapp(org, cert, "+8801711111111")

    assert "/public/certificates/456" in sent["data"]["MediaUrl"]
    assert "Download:" not in sent["data"]["Body"]
    assert "http" not in sent["data"]["Body"]


def test_whatsapp_document_send_rejects_localhost_base_url(db, monkeypatch):
    org = get_org_settings(db)
    org.wa_provider = "cloud"
    org.wa_token = "token"
    org.wa_phone_number_id = "phone-id"
    db.commit()
    cert = Certificate(id=789, certificate_no="ACME/2025-26/3", period="2025-26")
    monkeypatch.setattr(
        "app.services.dispatch.whatsapp_sender.get_settings",
        lambda: SimpleNamespace(
            public_base_url="http://localhost:8000",
            link_signing_secret="test-secret",
        ),
    )

    with pytest.raises(RuntimeError, match="public HTTPS PUBLIC_BASE_URL"):
        send_certificate_whatsapp(org, cert, "+8801711111111")


# ------------------------------------------------------ Rate hook -----------
def test_rate_hook_anomalies(db):
    r = rate_hook.apply_rate_updates(db, [
        rate_hook.RateUpdate(section="89", kind="tds", rate=0.05),
    ])
    assert r["applied"] == 1 and not r["anomalies"]
    # 5% -> 9% is an 80% relative jump -> flagged but applied.
    r = rate_hook.apply_rate_updates(db, [
        rate_hook.RateUpdate(section="89", kind="tds", rate=0.09,
                             effective_from=date(2026, 2, 1)),
    ])
    assert r["applied"] == 1
    assert any("Unexpected rate change" in m for m in r["anomalies"])
    # invalid kind / out-of-range rate rejected
    r = rate_hook.apply_rate_updates(db, [
        rate_hook.RateUpdate(section="89", kind="xyz", rate=0.05),
        rate_hook.RateUpdate(section="90", kind="tds", rate=5.0),
    ])
    assert r["applied"] == 0 and len(r["anomalies"]) == 2


# ------------------------------------------- Vendor onboarding (item 1) -----
def test_supplier_create_rejects_invalid_fields():
    with pytest.raises(ValidationError, match="12 digits"):
        SupplierCreate(company_id=1, name="Gamma", address="Dhaka", tin="12345", bin="B1",
                       email="a@b.com", whatsapp="+8801711111111")
    with pytest.raises(ValidationError, match="email"):
        SupplierCreate(company_id=1, name="Gamma", address="Dhaka", tin="111122223333", bin="B1",
                       email="not-an-email", whatsapp="+8801711111111")
    with pytest.raises(ValidationError, match="WhatsApp"):
        SupplierCreate(company_id=1, name="Gamma", address="Dhaka", tin="111122223333", bin="B1",
                       email="a@b.com", whatsapp="abc")
    with pytest.raises(ValidationError, match="required"):
        SupplierCreate(company_id=1, name="", address="Dhaka", tin="111122223333", bin="B1",
                       email="a@b.com", whatsapp="+8801711111111")


def test_supplier_onboarding_creates_then_upserts_by_tin(db, company):
    body = SupplierCreate(company_id=company.id, name="Gamma Traders", address="Dhaka",
                          tin="555566667777", bin="BIN001", email="gamma@example.com",
                          whatsapp="+8801711112222")
    sup = onboard_supplier(body, db)
    assert sup.tin == "555566667777" and sup.bin == "BIN001"
    emails = [c.value for c in sup.contacts if c.kind == ContactKind.EMAIL]
    assert emails == ["gamma@example.com"]

    # Re-onboarding the same (company, TIN) updates the existing supplier instead of duplicating.
    body2 = SupplierCreate(company_id=company.id, name="Gamma Traders Ltd", address="Chattogram",
                           tin="555566667777", bin="BIN002", email="new@example.com",
                           whatsapp="+8801711113333")
    sup2 = onboard_supplier(body2, db)
    assert sup2.id == sup.id
    assert sup2.name == "Gamma Traders Ltd" and sup2.bin == "BIN002"
    assert db.query(Supplier).filter_by(tin="555566667777").count() == 1


# --------------------------------------- Signature/seal split (item 8) ------
def test_missing_seal_anomaly_satisfied_by_company_seal(db, company, sample_xlsx):
    import_depot_workbook(db, sample_xlsx, "sample.xlsx", company.id)
    cert = generate_certificate(db, company.id, "111122223333", "2025-26")
    org = get_org_settings(db)

    codes = {a.code for a in check_certificate(db, cert, org)}
    assert "MISSING_SEAL" in codes  # no seal uploaded yet

    company.seal_data = _TINY_PNG
    db.commit()
    codes = {a.code for a in check_certificate(db, cert, org)}
    assert "MISSING_SEAL" not in codes


# ---------------------------------- Multiple named signatures (items 2/5) ---
def test_signature_defaults_enabled_and_renders_on_certificate(db, company, sample_xlsx):
    sig = Signature(company_id=company.id, name="Md. Rahim Uddin",
                    designation="Head of Tax & VAT", image_data=_TINY_PNG)
    db.add(sig)
    db.commit()
    assert sig.enabled is True  # new signatures default to enabled

    import_depot_workbook(db, sample_xlsx, "sample.xlsx", company.id)
    cert = generate_certificate(db, company.id, "444455556666", "2025-26")
    render_certificate_pdf(db, cert)  # must not raise with a signature row
    assert cert.pdf_data


def test_disabled_signature_excluded_from_enabled_query(db, company):
    """Every Signature flagged enabled renders on every certificate for the
    company — disabling one removes it from that set immediately."""
    shown = Signature(company_id=company.id, name="Shown", image_data=_TINY_PNG)
    hidden = Signature(company_id=company.id, name="Hidden", image_data=_TINY_PNG, enabled=False)
    db.add_all([shown, hidden])
    db.commit()

    visible = (
        db.query(Signature)
        .filter(Signature.company_id == company.id, Signature.enabled.is_(True))
        .order_by(Signature.name)
        .all()
    )
    assert [s.name for s in visible] == ["Shown"]


def test_delete_signature_route(db, company):
    from app.main import app
    from app.database import get_db

    sig = Signature(company_id=company.id, name="To Delete", image_data=_TINY_PNG)
    db.add(sig)
    db.commit()
    sig_id = sig.id

    app.dependency_overrides[get_db] = lambda: db
    try:
        client = TestClient(app)
        resp = client.delete(f"/api/companies/{company.id}/signatures/{sig_id}")
        assert resp.status_code == 200
        assert db.get(Signature, sig_id) is None
    finally:
        app.dependency_overrides.clear()


# ------------------------------------------------ Logo removal (item 1) -----
def test_company_logo_route_removed(db, company):
    """The per-company logo feature (upload + certificate rendering) has
    been removed entirely, not just hidden — the route no longer exists."""
    from app.main import app
    from app.database import get_db

    app.dependency_overrides[get_db] = lambda: db
    try:
        client = TestClient(app)
        resp = client.get(f"/api/companies/{company.id}/logo")
        assert resp.status_code == 404
    finally:
        app.dependency_overrides.clear()


def test_company_has_no_logo_path_column():
    assert not hasattr(Company, "logo_path")


# --------------------------------------- Pending groupings search (item 8) --
def test_pending_groupings_filters_match_certificate_search(db, company, sample_xlsx):
    """Pending now uses the same server-side filters as certificate search
    instead of a looser client-side substring match, so a query like "R.S"
    can no longer false-positive match an unrelated supplier."""
    import_depot_workbook(db, sample_xlsx, "sample.xlsx", company.id)

    by_name = list_groupings(db, company.id, supplier_name="Beta")
    assert {r["tin"] for r in by_name} == {"444455556666"}

    by_tin = list_groupings(db, company.id, tin="1111")
    assert {r["tin"] for r in by_tin} == {"111122223333"}

    by_bin = list_groupings(db, company.id, bin="nonexistent-bin")
    assert by_bin == []


# ------------------------------------ Configurable number format (item 9) ---
def test_number_format_default_matches_legacy_output(tmp_path):
    url = f"sqlite:///{tmp_path}/fmt.db"
    engine = create_engine(url, connect_args={"check_same_thread": False})
    Base.metadata.create_all(engine)
    s = sessionmaker(bind=engine, expire_on_commit=False)()
    company = Company(name="ACME", is_default=True)
    s.add(company)
    s.commit()
    cfg = get_numbering_config(s, company.id)
    cfg.company_token = "ACME"
    s.commit()
    n = allocate_certificate_number(s, company.id, "2025-26")
    assert n == "ACME/2025-26/1"


def test_number_format_supports_reordered_tokens(tmp_path):
    url = f"sqlite:///{tmp_path}/fmt2.db"
    engine = create_engine(url, connect_args={"check_same_thread": False})
    Base.metadata.create_all(engine)
    s = sessionmaker(bind=engine, expire_on_commit=False)()
    company = Company(name="ACME", is_default=True)
    s.add(company)
    s.commit()
    cfg = get_numbering_config(s, company.id)
    cfg.company_token = "ACME"
    cfg.number_format = "{FiscalYear}-{AutoNumber}-{CompanyName}"
    s.commit()
    n = allocate_certificate_number(s, company.id, "2025-26")
    assert n == "2025-26-1-ACME"


def test_number_format_independent_per_company(tmp_path):
    """Two companies with identical settings never collide (item 9 + full
    multi-tenant numbering)."""
    url = f"sqlite:///{tmp_path}/fmt3.db"
    engine = create_engine(url, connect_args={"check_same_thread": False})
    Base.metadata.create_all(engine)
    s = sessionmaker(bind=engine, expire_on_commit=False)()
    c1 = Company(name="ACME", is_default=True)
    c2 = Company(name="Beta")
    s.add_all([c1, c2])
    s.commit()
    for c in (c1, c2):
        cfg = get_numbering_config(s, c.id)
        cfg.company_token = c.name
        s.commit()

    n1a = allocate_certificate_number(s, c1.id, "2025-26")
    n2a = allocate_certificate_number(s, c2.id, "2025-26")
    n1b = allocate_certificate_number(s, c1.id, "2025-26")
    assert n1a == "ACME/2025-26/1"
    assert n2a == "Beta/2025-26/1"  # independent sequence, also starts at 1
    assert n1b == "ACME/2025-26/2"


# -------------------------------------------- Route ordering (item 11) ------
def test_export_route_not_shadowed_by_cert_id_route(db, company, sample_xlsx):
    """Regression test: /certificates/export must be registered before
    /certificates/{cert_id}, or FastAPI's registration-order path matching
    swallows "export" as a cert_id and 422s (found via live manual testing)."""
    from app.main import app
    from app.database import get_db

    import_depot_workbook(db, sample_xlsx, "sample.xlsx", company.id)
    generate_certificate(db, company.id, "444455556666", "2025-26")

    app.dependency_overrides[get_db] = lambda: db
    try:
        client = TestClient(app)
        resp = client.get("/api/certificates/export", params={"company_id": company.id})
        assert resp.status_code == 200, resp.text
        assert resp.headers["content-type"] == (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        # The ordinary get-by-id route must still resolve correctly too.
        cert_id = db.query(Certificate).first().id
        resp2 = client.get(f"/api/certificates/{cert_id}")
        assert resp2.status_code == 200
        assert resp2.json()["id"] == cert_id
    finally:
        app.dependency_overrides.clear()
